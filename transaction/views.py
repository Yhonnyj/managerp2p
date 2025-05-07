import openpyxl
import os
import pandas as pd

from .models import Transaction, Client, ClienteImagen
from .serializers import TransactionSerializer, ClienteSerializer  # ✅ Importamos ambos serializadores correctamente
from datetime import timedelta

from django.http import HttpResponse, JsonResponse
from django.conf import settings
from django.db.models import Count, Sum, F
from django_filters.rest_framework import DjangoFilterBackend
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.core.cache import cache
from django.db.models.functions import ExtractMonth, ExtractYear
from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework import viewsets, permissions, status, serializers
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.generics import CreateAPIView, ListAPIView, DestroyAPIView

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet

# DASHBOARD
@method_decorator(cache_page(300), name='dispatch')
class DashboardSummaryView(APIView):
    def get(self, request):
        hoy = timezone.now()
        mes_actual = hoy.month
        año_actual = hoy.year

        transacciones_mes = Transaction.objects.filter(
            date__month=mes_actual,
            date__year=año_actual
        )

        total = transacciones_mes.count()
        compras = transacciones_mes.filter(transaction_type="compra").count()
        ventas = transacciones_mes.filter(transaction_type="venta").count()
        ganancia = sum([t.profit for t in transacciones_mes])

        return Response({
            "transacciones_totales": total,
            "compras": compras,
            "ventas": ventas,
            "ganancia": ganancia,
        }, status=status.HTTP_200_OK)

# GRAFICO MENSUAL TRANSACCIONES
@method_decorator(cache_page(300), name='dispatch')
class MonthlyOperationsView(APIView):
    def get(self, request):
        operaciones_por_mes = (
            Transaction.objects
            .annotate(mes=ExtractMonth("date"))
            .values("mes")
            .annotate(total=Count("id"))
            .order_by("mes")
        )

        resumen = [0] * 12
        for item in operaciones_por_mes:
            mes_index = item["mes"] - 1
            resumen[mes_index] = item["total"]

        return Response({"series": resumen}, status=status.HTTP_200_OK)

# GRAFICO MENSUAL CLIENTES NUEVOS
@method_decorator(cache_page(300), name='dispatch')
class MonthlyNewClientsView(APIView):
    def get(self, request):
        clientes_por_mes = (
            Client.objects
            .annotate(mes=ExtractMonth("created_at"))
            .values("mes")
            .annotate(total=Count("id"))
            .order_by("mes")
        )

        resumen = [0] * 12
        for item in clientes_por_mes:
            mes_index = item["mes"] - 1
            resumen[mes_index] = item["total"]

        return Response({"series": resumen}, status=status.HTTP_200_OK)

# LISTA DE CLIENTES CON MÁS TRANSACCIONES
@method_decorator(cache_page(300), name='dispatch')
class TopClientesView(APIView):
    def get(self, request):
        top_clientes = (
            Transaction.objects
            .values("client", "client__nombre")
            .annotate(total=Count("id"))
            .order_by("-total")[:20]
        )

        data = [
            {
                "client": c["client__nombre"],
                "client_id": c["client"],  # 👈 esto es lo nuevo
                "total": c["total"]
            }
            for c in top_clientes
        ]

        return Response(data, status=status.HTTP_200_OK)


# ÚLTIMAS TRANSACCIONES
@method_decorator(cache_page(300), name='dispatch')
class UltimasTransaccionesView(APIView):
    def get(self, request):
        transacciones = (
            Transaction.objects
            .select_related("client")  # NO más "operador"
            .order_by("-date")[:20]
        )

        data = [
            {
                "id": t.id,
                "client_id": t.client.id,
                "client_name": t.client.nombre,
                "usd": float(t.usd),
                "usdt": float(t.usdt),
                "fee": float(t.fee),
                "profit": float(t.profit),
                "platform": t.platform or "-",
                "payment_method": t.payment_method or "N/A",
                "transaction_type": t.transaction_type,
                "date": t.date.strftime("%Y-%m-%d"),
            }
            for t in transacciones
        ]

        return Response(data, status=status.HTTP_200_OK)


# 🔶 Transacciones por Plataforma
@method_decorator(cache_page(300), name="dispatch")
class PlatformSummaryView(APIView):
    def get(self, request):
        resumen = (
            Transaction.objects
            .values("platform")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        data = [{"name": item["platform"], "value": item["total"]} for item in resumen]
        return Response(data, status=status.HTTP_200_OK)

# 🔶 Transacciones por Método de Pago
@method_decorator(cache_page(300), name="dispatch")
class PaymentMethodSummaryView(APIView):
    def get(self, request):
        resumen = (
            Transaction.objects
            .values("payment_method")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        data = [{"name": item["payment_method"], "value": item["total"]} for item in resumen]
        return Response(data, status=status.HTTP_200_OK)

# 🔶 Compras vs Ventas (por tipo de transacción)
@method_decorator(cache_page(300), name="dispatch")
class TipoTransaccionSummaryView(APIView):
    def get(self, request):
        compras = Transaction.objects.filter(transaction_type="compra").count()
        ventas = Transaction.objects.filter(transaction_type="venta").count()
        data = [
            {"name": "Compras", "value": compras},
            {"name": "Ventas", "value": ventas}
        ]
        return Response(data, status=status.HTTP_200_OK)

# 🔶 Reporte filtrado (no se cachea porque depende de query params)
class FilteredDashboardSummaryView(APIView):
    def get(self, request):
        filtro = request.GET.get("tipo", "mes")  # "semana" o "mes"
        mes = request.GET.get("mes")  # mes numérico
        anio = request.GET.get("anio")  # año numérico

        ahora = timezone.now()

        if filtro == "semana":
            fecha_inicio = ahora - timedelta(days=7)
            transacciones = Transaction.objects.filter(date__gte=fecha_inicio)
        else:
            if not mes or not anio:
                return Response({"error": "Mes y año son requeridos para el filtro mensual."}, status=400)
            transacciones = Transaction.objects.filter(date__month=mes, date__year=anio)

        total = transacciones.count()
        ganancia = sum([t.profit for t in transacciones])

        compras = transacciones.filter(transaction_type="compra")
        ventas = transacciones.filter(transaction_type="venta")

        compra_usdt = compras.aggregate(usdt=Sum("usdt"))['usdt'] or 0
        compra_usd = compras.aggregate(usd=Sum("usd"))['usd'] or 0
        venta_usdt = ventas.aggregate(usdt=Sum("usdt"))['usdt'] or 0
        venta_usd = ventas.aggregate(usd=Sum("usd"))['usd'] or 0

        return Response({
            "transacciones_totales": total,
            "ganancia": float(ganancia),
            "compras": {
                "usdt": float(compra_usdt),
                "usd": float(compra_usd)
            },
            "ventas": {
                "usdt": float(venta_usdt),
                "usd": float(venta_usd)
            }
        }, status=status.HTTP_200_OK)

# ✅ VISTASET de transacciones con filtro por cliente (ideal para paginación en frontend)
class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.select_related("client").all().order_by("-date")
    serializer_class = TransactionSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['client']

# ✅ Lista de clientes con total de transacciones (cacheado 5 min)
@cache_page(300)
@api_view(["GET"])
def get_unique_clients(request):
    clients = Client.objects.annotate(total_transacciones=Count("transaction"))
    serializer = ClienteSerializer(clients, many=True)
    return JsonResponse(serializer.data, safe=False)

# ✅ Todos los clientes (cacheado 5 min)
@cache_page(300)
@api_view(["GET"])
def get_clients(request):
    clients = Client.objects.all()
    serializer = ClienteSerializer(clients, many=True)
    return JsonResponse(serializer.data, safe=False)

# ✅ Eliminar cliente (no se cachea)
@api_view(["DELETE"])
def delete_client(request, client_id):
    try:
        client = Client.objects.get(id=client_id)
        client.delete()
        return JsonResponse({"message": "Cliente eliminado correctamente"}, status=200)
    except Client.DoesNotExist:
        return JsonResponse({"error": "Cliente no encontrado"}, status=404)

# ✅ Obtener cliente por ID (cacheado 5 min)
@cache_page(300)
@api_view(["GET"])
def get_client_by_id(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    serializer = ClienteSerializer(client)
    return JsonResponse(serializer.data, safe=False)

# ✅ Buscar cliente por nombre exacto (cacheado 5 min)
@cache_page(300)
@api_view(["GET"])
def get_client_by_name(request, client_name):
    client = get_object_or_404(Client, nombre__iexact=client_name)
    serializer = ClienteSerializer(client)
    return JsonResponse(serializer.data, safe=False)

# ✅ Lista completa de clientes para modales (cacheado 1 hora)
class ClienteFullListView(APIView):
    def get(self, request):
        cache_key = "clientes_full"
        data = cache.get(cache_key)
        if not data:
            clientes = Client.objects.all()
            serializer = ClienteSerializer(clientes, many=True)
            data = serializer.data
            cache.set(cache_key, data, timeout=3600)
        return Response(data, status=status.HTTP_200_OK)

# ✅ Resumen de clientes con total de transacciones agrupadas (cacheado 30 min)
class ClienteResumenView(APIView):
    def get(self, request):
        cache_key = "clientes_resumen"
        data = cache.get(cache_key)
        if not data:
            clientes = (
                Client.objects
                .annotate(total_transacciones=Count("transaction"))
                .values("id", "nombre", "total_transacciones")
                .order_by("id")
            )
            data = list(clientes)
            cache.set(cache_key, data, timeout=1800)
        return Response(data, status=status.HTTP_200_OK)

# ✅ Vista para crear nuevos clientes
class ClienteCreateView(CreateAPIView):
    queryset = Client.objects.all()
    serializer_class = ClienteSerializer

    def perform_create(self, serializer):
        cliente = serializer.save()
        # 🔄 Invalida caché relevante al crear nuevo cliente
        cache.delete("clientes_full")
        cache.delete("clientes_resumen")

# ✅ Vista para actualizar datos del cliente por ID
class ClienteUpdateView(APIView):
    def put(self, request, client_id):
        try:
            cliente = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            return Response({"error": "Cliente no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        update_data = {
            key: request.data.get(key)
            for key in ["nombre", "email", "telefono", "pais", "direccion"]
            if request.data.get(key) is not None
        }

        for key, value in update_data.items():
            setattr(cliente, key, value)
        cliente.save()

        # 🔄 Invalida caché relevante al actualizar
        cache.delete("clientes_full")
        cache.delete("clientes_resumen")

        return Response({"message": "Cliente actualizado correctamente"}, status=status.HTTP_200_OK)


# ✅ Reporte completo optimizado con Redis por filtro
class ReporteCompletoView(APIView):
    def get(self, request):
        filtro = request.GET.get("tipo", "mes")
        mes = request.GET.get("mes")
        anio = request.GET.get("anio")

        cache_key = f"reporte_completo:{filtro}:{mes or 'na'}:{anio or 'na'}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data, status=status.HTTP_200_OK)

        ahora = timezone.now()

        if filtro == "semana":
            fecha_inicio = ahora - timedelta(days=7)
            transacciones = Transaction.objects.filter(date__gte=fecha_inicio)
        else:
            if not mes or not anio:
                return Response({"error": "Mes y año son requeridos para el filtro mensual."}, status=400)
            transacciones = Transaction.objects.filter(date__month=mes, date__year=anio)

        # 🟨 Resumen
        total = transacciones.count()
        ganancia = sum([t.profit for t in transacciones])
        compras = transacciones.filter(transaction_type="compra")
        ventas = transacciones.filter(transaction_type="venta")
        resumen = {
            "transacciones_totales": total,
            "ganancia": float(ganancia),
            "compras": {
                "usdt": float(compras.aggregate(usdt=Sum("usdt"))['usdt'] or 0),
                "usd": float(compras.aggregate(usd=Sum("usd"))['usd'] or 0)
            },
            "ventas": {
                "usdt": float(ventas.aggregate(usdt=Sum("usdt"))['usdt'] or 0),
                "usd": float(ventas.aggregate(usd=Sum("usd"))['usd'] or 0)
            }
        }

        # 🟩 Plataformas
        plataformas = (
            transacciones.values("platform")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        plataformas_data = [{"name": p["platform"], "value": p["total"]} for p in plataformas]

        # 🟦 Métodos de Pago
        metodos = (
            transacciones.values("payment_method")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        metodos_pago_data = [{"name": m["payment_method"], "value": m["total"]} for m in metodos]

        # 🟥 Tipos de transacción
        tipos_data = [
            {"name": "Compras", "value": compras.count()},
            {"name": "Ventas", "value": ventas.count()}
        ]

        # 🟪 Operaciones por mes (últimos 12 meses)
        meses = Transaction.objects.annotate(
            mes=ExtractMonth("date"),
            anio=ExtractYear("date")
        ).values("mes", "anio").annotate(total=Count("id")).order_by("anio", "mes")
        operaciones_mensuales = [
            {"name": f"{m['mes']:02d}-{m['anio']}", "value": m["total"]}
            for m in meses
        ]

        resultado = {
            "resumen": resumen,
            "plataformas": plataformas_data,
            "metodos_pago": metodos_pago_data,
            "tipos_transaccion": tipos_data,
            "operaciones_mensuales": operaciones_mensuales
        }

        # 🧠 Guardar en caché 30 minutos
        cache.set(cache_key, resultado, timeout=1800)

        return Response(resultado, status=status.HTTP_200_OK)

#EXPORTAR EXCEL

# ✅ Exportar transacciones a Excel (sin caché porque genera archivo)
@api_view(["GET"])
def export_transactions_to_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Transacciones"

    headers = ["Sell Price", "USDT", "USD", "Fee", "Profit", "Client", "Date", "Platform"]
    sheet.append(headers)

    transactions = Transaction.objects.all()
    for t in transactions:
        fee_amount = (t.usdt * t.fee) / 100
        profit = round(
            (t.usdt - t.usd - fee_amount)
            if t.transaction_type == "compra"
            else (t.usd - t.usdt - fee_amount),
            2
        )
        sell_price = round(t.usdt / t.usd, 2) if t.transaction_type == "compra" else round(t.usd / t.usdt, 2)

        sheet.append([
            sell_price,
            float(t.usdt),
            float(t.usd),
            round(fee_amount, 2),
            profit,
            t.client.nombre,
            t.date.strftime('%Y-%m-%d'),
            t.platform
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="transacciones.xlsx"'
    workbook.save(response)
    return response

# ✅ Exportar clientes a Excel
@api_view(["GET"])
def export_clients_to_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Clientes"

    headers = ["ID", "Nombre", "Correo", "Teléfono", "País", "Dirección"]
    sheet.append(headers)

    clients = Client.objects.all()
    for client in clients:
        sheet.append([
            client.id,
            client.nombre,
            client.email if client.email else "N/A",
            client.telefono if client.telefono else "N/A",
            client.pais if client.pais else "N/A",
            client.direccion if client.direccion else "N/A"
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="clientes.xlsx"'
    workbook.save(response)
    return response

# ✅ Exportar transacciones de un cliente a Excel
@api_view(["GET"])
def export_client_transactions_to_excel(request, client_id):
    # 🔍 Filtrar transacciones por el cliente
    transactions = Transaction.objects.filter(client_id=client_id)

    if not transactions.exists():
        return HttpResponse("No hay transacciones para este cliente.", status=404)

    # 📂 Crear archivo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Transacciones"

    # ✅ Encabezados exactos que solicitaste
    headers = ["USDT", "USD", "Profit", "Pago", "Plataforma", "Fecha"]
    sheet.append(headers)

    # 📌 Agregar datos de transacciones
    for transaction in transactions:
        sheet.append([
            transaction.usdt,
            transaction.usd,
            transaction.profit,
            transaction.payment_method if transaction.payment_method else "N/A",
            transaction.platform,
            transaction.date.strftime('%Y-%m-%d')
        ])

    # 📩 Enviar archivo como respuesta
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="transacciones_cliente_{client_id}.xlsx"'
    workbook.save(response)
    return response

#EXPORTAR PDF

# ✅ Exportar transacciones a PDF (NO SE CACHEA)
@api_view(["GET"])
def export_transactions_to_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_transacciones.pdf"'

    pdf = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    logo_path = os.path.join(settings.BASE_DIR, "managerp2p/static/logo.png")
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=120, height=120)
        elements.append(logo)

    title = Paragraph("<b>📄 Reporte de Transacciones</b>", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))

    transactions = Transaction.objects.all()
    total_transactions = transactions.count()
    total_usdt = sum(t.usdt for t in transactions)
    total_usd = sum(t.usd for t in transactions)

    total_profit = sum(
        round(
            (t.usdt - t.usd - (t.usdt * t.fee) / 100)
            if t.transaction_type == "compra"
            else (t.usd - t.usdt - (t.usdt * t.fee) / 100),
            2
        ) for t in transactions
    )

    resumen_data = [
        ["Total de Transacciones:", total_transactions],
        ["Total USDT:", f"{total_usdt:,.2f}"],
        ["Total USD:", f"{total_usd:,.2f}"],
        ["Ganancia Total (Profit):", f"{total_profit:,.2f}"]
    ]

    resumen_table = Table(resumen_data, colWidths=[250, 150])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5)
    ]))

    elements.append(resumen_table)
    elements.append(Spacer(1, 12))

    headers = ["Sell Price", "USDT", "USD", "Fee", "Profit", "Client", "Date", "Platform"]
    data = [headers]

    for t in transactions:
        fee_amount = (t.usdt * t.fee) / 100
        profit = round(
            (t.usdt - t.usd - fee_amount)
            if t.transaction_type == "compra"
            else (t.usd - t.usdt - fee_amount),
            2
        )
        sell_price = round(t.usdt / t.usd, 2) if t.transaction_type == "compra" else round(t.usd / t.usdt, 2)

        data.append([
            f"{sell_price:,.2f}",
            f"{t.usdt:,.2f}",
            f"{t.usd:,.2f}",
            f"{fee_amount:,.2f}",
            f"{profit:,.2f}",
            t.client.nombre,
            t.date.strftime('%Y-%m-%d'),
            t.platform
        ])

    table = Table(data, colWidths=[80, 80, 80, 80, 80, 120, 100, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
    ]))

    elements.append(table)
    pdf.build(elements)

    return response

# ✅ Exportar clientes a PDF
@api_view(["GET"])
def export_clients_to_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="clientes.pdf"'

    pdf = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # ✅ Título del PDF
    title = Paragraph("<b>📄 Lista de Clientes</b>", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # ✅ Datos de los clientes
    clients = Client.objects.all()
    data = [["ID", "Nombre", "Correo", "Teléfono", "País", "Dirección"]]

    for client in clients:
        data.append([
            client.id,
            client.nombre,
            client.email if client.email else "N/A",
            client.telefono if client.telefono else "N/A",
            client.pais if client.pais else "N/A",
            client.direccion if client.direccion else "N/A"
        ])

    table = Table(data, colWidths=[30, 120, 120, 80, 80, 150])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
    ]))

    elements.append(table)
    pdf.build(elements)

    return response


# ✅ Exportar transacciones de un cliente específico a PDF
@api_view(["GET"])
def export_client_transactions_to_pdf(request, client_id):
    # 📄 Configurar el documento PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="transacciones_cliente_{client_id}.pdf"'

    pdf = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # 📌 Título del PDF
    title = Paragraph(f"<b>📄 Historial de Transacciones - Cliente {client_id}</b>", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))

   
    # ✅ Obtener transacciones del cliente específico
    transactions = Transaction.objects.filter(client__id=client_id)

    if not transactions.exists():
        no_data = Paragraph("<b>No hay transacciones para este cliente.</b>", styles["Normal"])
        elements.append(no_data)
        pdf.build(elements)
        return response

    # ✅ Encabezados correctos
    data = [["USDT", "USD", "Profit", "Pago", "Plataforma", "Fecha"]]

    # 📌 Agregar datos de transacciones
    for transaction in transactions:
        data.append([
            f"{transaction.usdt}",
            f"{transaction.usd}",
            f"{transaction.profit}",
            transaction.payment_method if transaction.payment_method else "N/A",
            transaction.platform,
            transaction.date.strftime('%Y-%m-%d')
        ])

    # 📊 Configuración de la tabla
    table = Table(data, colWidths=[60, 60, 60, 80, 80, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
    ]))

    elements.append(table)
    pdf.build(elements)

    return response


# ✅ Importar transacciones desde Excel (NO SE CACHEA)
@api_view(["POST"])
@parser_classes([MultiPartParser])
def import_transactions_from_excel(request):
    if "file" not in request.FILES:
        return JsonResponse({"error": "No se envió ningún archivo."}, status=400)

    file = request.FILES["file"]

    try:
        df = pd.read_excel(file)
        df.columns = df.columns.str.strip().str.lower()

        column_mapping = {
            "usdt": "usdt",
            "usd": "usd",
            "fee": "fee",
            "cliente": "client",
            "fecha": "date",
            "plataforma": "platform",
            "transaction_type": "transaction_type"
        }

        df.rename(columns=column_mapping, inplace=True)

        required_columns = set(column_mapping.values())
        missing_columns = required_columns - set(df.columns)
        if missing_columns:
            return JsonResponse({"error": f"Faltan columnas: {', '.join(missing_columns)}"}, status=400)

        df["date"] = pd.to_datetime(df["date"], errors="coerce")

        for _, row in df.iterrows():
            client, _ = Client.objects.get_or_create(nombre=row["client"])
            tipo = str(row["transaction_type"]).strip().lower()
            if tipo not in ["compra", "venta"]:
                tipo = "compra"

            Transaction.objects.create(
                usdt=row["usdt"],
                usd=row["usd"],
                fee=row["fee"],
                client=client,
                date=row["date"],
                platform=row["platform"],
                transaction_type=tipo
            )

        return JsonResponse({
            "message": "Importación exitosa",
            "transactions": df.to_dict(orient="records")
        }, status=201)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    

from transaction.tasks import export_transactions_excel_task

@api_view(["GET"])
def generar_excel_transacciones(request):
    task = export_transactions_excel_task.delay()
    return Response({
        "message": "📝 Exportación iniciada",
        "task_id": str(task.id)
    })


#IMAGENES
class ClienteImagenSerializer(serializers.ModelSerializer):
    imagen = serializers.SerializerMethodField()

    class Meta:
        model = ClienteImagen
        fields = ["id", "titulo", "imagen", "creado"]

    def get_imagen(self, obj):
        return obj.imagen.url if obj.imagen else None


class ImagenesClienteView(ListAPIView):
    serializer_class = ClienteImagenSerializer

    def get_queryset(self):
        client_id = self.kwargs["client_id"]
        return ClienteImagen.objects.filter(cliente_id=client_id).order_by("-creado")


class SubirImagenClienteView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, client_id):
        try:
            cliente = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            return Response({"error": "Cliente no encontrado."}, status=404)

        imagen = request.FILES.get("imagen")
        titulo = request.data.get("titulo", "")

        if not imagen:
            return Response({"error": "Imagen no enviada."}, status=400)

        img = ClienteImagen.objects.create(cliente=cliente, imagen=imagen, titulo=titulo)
        return Response({
            "id": img.id,
            "url": img.imagen.url,
            "titulo": img.titulo,
            "creado": img.creado
        }, status=201)


class EliminarImagenView(DestroyAPIView):
    queryset = ClienteImagen.objects.all()
    lookup_field = "pk"

